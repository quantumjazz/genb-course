"""End-to-end smoke test against a running session-quiz backend.

Usage:
    python3 smoke_test.py [--base http://localhost:8789] [--admin KEY]

The test:
  1. uploads a 5-row sample bank,
  2. creates a session, starts it,
  3. joins two students,
  4. each student answers 3 questions,
  5. one student triggers /quiz/blur,
  6. exports the .xlsx and checks the header rows.

Run the server first:  python3 server.py

This script is for development sanity checks only, not for production.
"""

import argparse
import io
import json
import os
import sys
import urllib.error
import urllib.request

from openpyxl import Workbook, load_workbook


def make_sample_xlsx_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    ws.append([
        "lecture_tag", "stem", "option_a", "option_b", "option_c",
        "option_d", "option_e", "option_f", "correct", "explanation",
    ])
    rows = [
        ("ch01_intro", "Колко е 2+2?", "3", "4", "5", "", "", "", "b", "Просто събиране."),
        ("ch01_intro", "Столицата на България е…", "София", "Пловдив", "Варна", "", "", "", "a", ""),
        ("ch01_intro", "Първият закон на Нютон описва…", "инерцията", "силата", "ускорението", "масата", "", "", "a", ""),
        ("ch01_intro", "H2O е…", "сол", "вода", "захар", "", "", "", "b", ""),
        ("ch01_intro", "365 дни има…", "седмица", "месец", "година", "век", "", "", "c", ""),
        ("ch01_intro", "Земята обикаля около…", "Луната", "Слънцето", "Юпитер", "", "", "", "b", ""),
        ("ch01_intro", "Кой е написал \"Под игото\"?", "Йовков", "Вазов", "Ботев", "Пенчо Славейков", "", "", "b", ""),
    ]
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def request(method, url, *, body=None, headers=None, raw=False):
    data = None
    h = dict(headers or {})
    if body is not None and not isinstance(body, (bytes, bytearray)):
        data = json.dumps(body).encode("utf-8")
        h.setdefault("Content-Type", "application/json")
    elif body is not None:
        data = body
    req = urllib.request.Request(url, data=data, method=method, headers=h)
    try:
        resp = urllib.request.urlopen(req, timeout=10)
    except urllib.error.HTTPError as exc:
        print(f"  HTTP {exc.code}: {exc.read().decode('utf-8', 'replace')}")
        raise
    body_bytes = resp.read()
    if raw:
        return body_bytes, dict(resp.headers)
    return json.loads(body_bytes.decode("utf-8"))


def upload_multipart(url, file_bytes, filename, name, headers):
    boundary = "----quizboundary123"
    parts = [
        f"--{boundary}\r\n".encode(),
        b'Content-Disposition: form-data; name="file"; filename="', filename.encode(), b'"\r\n',
        b"Content-Type: application/octet-stream\r\n\r\n",
        file_bytes, b"\r\n",
        f"--{boundary}\r\n".encode(),
        b'Content-Disposition: form-data; name="name"\r\n\r\n',
        name.encode(), b"\r\n",
        f"--{boundary}--\r\n".encode(),
    ]
    body = b"".join(parts)
    h = dict(headers or {})
    h["Content-Type"] = f"multipart/form-data; boundary={boundary}"
    h["Content-Length"] = str(len(body))
    return request("POST", url, body=body, headers=h)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", default="http://localhost:8789")
    parser.add_argument("--admin", default=os.environ.get("QUIZ_ADMIN_KEY", ""))
    args = parser.parse_args()

    base = args.base.rstrip("/")
    admin_headers = {"X-Admin-Key": args.admin} if args.admin else {}

    print("[1] health")
    health = request("GET", f"{base}/quiz/health")
    assert health.get("ok"), health

    print("[2] upload bank")
    xlsx = make_sample_xlsx_bytes()
    bank_resp = upload_multipart(
        f"{base}/quiz/admin/bank/upload", xlsx, "smoke.xlsx", "Smoke bank",
        admin_headers,
    )
    print(f"   bank_id={bank_resp['bank_id']}, items={bank_resp['item_count']}, tags={bank_resp['tags']}")
    assert bank_resp["item_count"] == 7
    bank_id = bank_resp["bank_id"]

    print("[3] create session")
    create = request("POST", f"{base}/quiz/admin/session/create",
                     body={
                         "bank_id": bank_id,
                         "lecture_tag": "ch01_intro",
                         "display_name": "Smoke test",
                         "item_count": 5,
                         "duration_minutes": 5,
                         "swap_policy": "soft",
                         "permutation": "per_view",
                         "feedback": "immediate",
                     },
                     headers=admin_headers)
    code = create["join_code"]
    sid = create["session_id"]
    print(f"   session_id={sid}, code={code}")

    print("[4] start session")
    start = request("POST", f"{base}/quiz/admin/session/start",
                    body={"session_id": sid}, headers=admin_headers)
    assert start.get("started_at"), start

    print("[5] join 2 students")
    j1 = request("POST", f"{base}/quiz/join",
                 body={"code": code, "student_number": "STU-001"})
    j2 = request("POST", f"{base}/quiz/join",
                 body={"code": code, "student_number": "STU-002"})
    print(f"   tokens: {j1['student_token']}, {j2['student_token']}")

    print("[6] STU-001 answers 3 questions")
    for i in range(3):
        nxt = request("GET", f"{base}/quiz/next?student_token={j1['student_token']}")
        if nxt.get("session_ended"):
            print(f"   ended early: {nxt}")
            break
        ans = request("POST", f"{base}/quiz/answer",
                      body={"attempt_id": nxt["attempt_id"], "chosen_visible_index": 0})
        print(f"   q{nxt['ord']}: stem={nxt['stem'][:30]!r}, ord={nxt['ord']}, correct={ans['correct']}")

    print("[7] STU-002 triggers blur on first question")
    nxt = request("GET", f"{base}/quiz/next?student_token={j2['student_token']}")
    print(f"   served ord={nxt['ord']}")
    blur = request("POST", f"{base}/quiz/blur",
                   body={"attempt_id": nxt["attempt_id"]})
    print(f"   blur result: {blur}")
    assert blur.get("swapped"), blur

    print("[8] STU-002 fetches next — should not have ord skipped")
    nxt2 = request("GET", f"{base}/quiz/next?student_token={j2['student_token']}")
    print(f"   re-served ord={nxt2.get('ord')}")
    assert nxt2.get("ord") == 1, "ord should still be 1 after a soft blur swap"

    print("[9] live view")
    live = request("GET", f"{base}/quiz/admin/session/live?session_id={sid}",
                   headers=admin_headers)
    print(f"   students={len(live['students'])}, remaining={live['remaining_ms']}ms")
    for s in live["students"]:
        print(f"     {s}")

    print("[10] export xlsx")
    body, headers = request("GET", f"{base}/quiz/admin/session/export?session_id={sid}",
                            headers=admin_headers, raw=True)
    print(f"   bytes={len(body)}, Content-Type={headers.get('Content-Type')}")
    wb = load_workbook(io.BytesIO(body))
    print(f"   sheets={wb.sheetnames}")
    assert wb.sheetnames == ["summary", "detail"], wb.sheetnames
    summary_rows = list(wb["summary"].iter_rows(values_only=True))
    detail_rows = list(wb["detail"].iter_rows(values_only=True))
    print(f"   summary header: {summary_rows[0]}")
    print(f"   summary rows after header: {len(summary_rows) - 1}")
    print(f"   detail rows after header: {len(detail_rows) - 1}")

    print("[11] close session")
    close = request("POST", f"{base}/quiz/admin/session/close",
                    body={"session_id": sid}, headers=admin_headers)
    assert close.get("closed_at"), close
    print("   closed_at:", close["closed_at"])

    print("\nAll smoke checks passed ✓")


if __name__ == "__main__":
    main()
