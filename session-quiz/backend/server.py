#!/usr/bin/env python3
"""Session-quiz classroom MCQ backend.

Serves student and instructor dashboards for the session-quiz app
described in ``../session-quiz-spec.md``. Stdlib HTTP server, SQLite,
per-session HMAC student pseudonymization, openpyxl for bank I/O and
Excel export, qrcode for join-code QR.
"""

import argparse
import base64
import hashlib
import hmac
import io
import json
import os
import random
import re
import secrets
import threading
import time
import uuid
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, urlparse

from openpyxl import Workbook

import db as dbmod
import bank_io
import qr as qrmod


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SESSION_CODE_CHARS = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"  # ambiguous chars removed
SESSION_CODE_LEN = 6
PSEUDONYM_KEY_BYTES = 32
TOKEN_LEN_CHARS = 12

DEFAULTS = {
    "item_count": 30,
    "duration_minutes": 60,
    "swap_policy": "soft",
    "permutation": "per_view",
    "feedback": "immediate",
    "exhaustion_policy": "end_session",
    "security_mode": "standard",
}

VALID = {
    "swap_policy": {"soft", "hard"},
    "permutation": {"per_view", "per_student"},
    "feedback": {"immediate", "end_of_session"},
    "exhaustion_policy": {"end_session", "recycle"},
    "security_mode": {"standard", "strict"},
}

ITEM_COUNT_RANGE = (5, 100)
DURATION_RANGE = (5, 180)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def now_ms():
    return int(time.time() * 1000)


def new_id():
    return uuid.uuid4().hex


def json_dumps(value):
    return json.dumps(value, ensure_ascii=False)


def parse_json(body):
    if not body:
        return {}
    return json.loads(body.decode("utf-8"))


def make_session_code():
    return "".join(secrets.choice(SESSION_CODE_CHARS) for _ in range(SESSION_CODE_LEN))


def derive_student_token(pseudonym_key, student_number):
    digest = hmac.new(
        pseudonym_key, student_number.encode("utf-8"), hashlib.sha256
    ).digest()
    b32 = base64.b32encode(digest).decode("ascii").rstrip("=")
    return b32[:TOKEN_LEN_CHARS]


_STUDENT_NUMBER_RE = re.compile(r"^[A-Z0-9][A-Z0-9\-]{1,31}$")


def normalize_student_number(value):
    if not isinstance(value, str):
        raise ValueError("Student number must be text.")
    cleaned = value.strip().upper().replace(" ", "")
    if not cleaned:
        raise ValueError("Student number cannot be empty.")
    if not _STUDENT_NUMBER_RE.match(cleaned):
        raise ValueError(
            "Student number must be 2–32 characters, letters/digits/hyphens."
        )
    return cleaned


def clamp_int(value, name, lo, hi, default):
    if value is None or value == "":
        return default
    try:
        n = int(value)
    except (TypeError, ValueError):
        raise ValueError(f"{name} must be a whole number.")
    if n < lo or n > hi:
        raise ValueError(f"{name} must be between {lo} and {hi}.")
    return n


def pick_enum(value, name, choices, default):
    if value is None or value == "":
        return default
    if value not in choices:
        raise ValueError(
            f"{name} must be one of: {', '.join(sorted(choices))}."
        )
    return value


def _clean_tag_list(values):
    tags = []
    seen = set()
    for value in values:
        tag = str(value or "").strip()
        if not tag or tag in seen:
            continue
        tags.append(tag)
        seen.add(tag)
    return tags


def parse_session_tags(value):
    """Return the tag list stored on quiz_session.lecture_tag.

    Older sessions store a single plain tag. New multi-tag sessions store a
    JSON list in the same column so the schema remains compatible.
    """
    text = str(value or "").strip()
    if not text:
        return []
    if text.startswith("["):
        try:
            loaded = json.loads(text)
        except json.JSONDecodeError:
            loaded = None
        if isinstance(loaded, list):
            return _clean_tag_list(loaded)
    return _clean_tag_list(text.split(","))


def session_tags(session_row):
    return parse_session_tags(session_row["lecture_tag"])


def session_tag_label(tags):
    return " + ".join(tags)


def encode_session_tags(tags):
    return tags[0] if len(tags) == 1 else json_dumps(tags)


def requested_tags(payload):
    raw = payload.get("lecture_tags")
    if isinstance(raw, list):
        return _clean_tag_list(raw)
    if isinstance(raw, str) and raw.strip():
        text = raw.strip()
        if text.startswith("["):
            try:
                loaded = json.loads(text)
            except json.JSONDecodeError:
                loaded = None
            if isinstance(loaded, list):
                return _clean_tag_list(loaded)
        return _clean_tag_list(text.split(","))
    return _clean_tag_list([payload.get("lecture_tag")])


# ---------------------------------------------------------------------------
# Rules text (Bulgarian, derived from session params)
# ---------------------------------------------------------------------------


def rules_text(session_row):
    item_count = session_row["item_count"]
    minutes = session_row["duration_minutes"]
    swap = session_row["swap_policy"]
    feedback = session_row["feedback"]

    parts = [f"{item_count} въпроса, {minutes} минути."]
    if swap == "soft":
        parts.append(
            "Ако напуснете страницата или превключите приложение, "
            "въпросът се сменя с друг."
        )
    else:
        parts.append(
            "Ако напуснете страницата или превключите приложение, "
            "сесията ви приключва."
        )
    if feedback == "immediate":
        parts.append("След всеки отговор виждате дали е верен.")
    else:
        parts.append("Резултатите се показват в края на теста.")
    if session_row["security_mode"] == "strict":
        parts.append(
            "Тестът изисква режим на цял екран; напускане на цял екран "
            "или промяна на прозореца сменя текущия въпрос."
        )
    return " ".join(parts)


# ---------------------------------------------------------------------------
# Session lifecycle helpers
# ---------------------------------------------------------------------------


def end_student(conn, session_id, student_token, reason):
    conn.execute(
        """UPDATE quiz_student
               SET ended_at = ?, end_reason = ?
             WHERE session_id = ? AND student_token = ?
               AND ended_at IS NULL""",
        (now_ms(), reason, session_id, student_token),
    )


def ending_reason_for_student(conn, session_row, student):
    """Return an end reason if this student can no longer answer."""
    if student["ended_at"]:
        return student["end_reason"] or "instructor_closed"
    if session_row["status"] == "closed":
        end_student(conn, session_row["id"], student["student_token"], "instructor_closed")
        return "instructor_closed"
    if session_row["started_at"] and compute_remaining_ms(session_row) <= 0:
        end_student(conn, session_row["id"], student["student_token"], "time_up")
        return "time_up"
    return None


def student_row(conn, session_id, student_token):
    return conn.execute(
        """SELECT * FROM quiz_student
                WHERE session_id = ? AND student_token = ?""",
        (session_id, student_token),
    ).fetchone()


def session_by_code(conn, code):
    return conn.execute(
        "SELECT * FROM quiz_session WHERE join_code = ?", (code.upper(),)
    ).fetchone()


def session_by_id(conn, session_id):
    return conn.execute(
        "SELECT * FROM quiz_session WHERE id = ?", (session_id,)
    ).fetchone()


def count_answered(conn, session_id, student_token):
    return conn.execute(
        """SELECT COUNT(*) AS n FROM quiz_attempt
            WHERE session_id = ? AND student_token = ?
              AND swapped = 0 AND submitted_at IS NOT NULL""",
        (session_id, student_token),
    ).fetchone()["n"]


def count_correct(conn, session_id, student_token):
    return conn.execute(
        """SELECT COUNT(*) AS n FROM quiz_attempt
            WHERE session_id = ? AND student_token = ?
              AND swapped = 0 AND correct = 1""",
        (session_id, student_token),
    ).fetchone()["n"]


def count_ord(conn, session_id, student_token):
    """Highest ord assigned so far (counted items only)."""
    row = conn.execute(
        """SELECT COALESCE(MAX(ord), 0) AS m FROM quiz_attempt
            WHERE session_id = ? AND student_token = ? AND ord IS NOT NULL""",
        (session_id, student_token),
    ).fetchone()
    return row["m"]


def count_blur(conn, session_id, student_token):
    return conn.execute(
        """SELECT COUNT(*) AS n FROM quiz_attempt
            WHERE session_id = ? AND student_token = ? AND swapped = 1""",
        (session_id, student_token),
    ).fetchone()["n"]


def count_incidents(conn, session_id, student_token):
    return conn.execute(
        """SELECT COUNT(*) AS n FROM quiz_incident
            WHERE session_id = ? AND student_token = ?""",
        (session_id, student_token),
    ).fetchone()["n"]


def seen_item_ids(conn, session_id, student_token):
    rows = conn.execute(
        """SELECT DISTINCT bank_item_id FROM quiz_attempt
            WHERE session_id = ? AND student_token = ?""",
        (session_id, student_token),
    ).fetchall()
    return {r["bank_item_id"] for r in rows}


def compute_remaining_ms(session_row):
    started_at = session_row["started_at"]
    if not started_at:
        return session_row["duration_minutes"] * 60_000
    end_ms = started_at + session_row["duration_minutes"] * 60_000
    return max(0, end_ms - now_ms())


# ---------------------------------------------------------------------------
# Quiz engine: serve next item / grade / blur
# ---------------------------------------------------------------------------


def open_attempt(conn, session_id, student_token):
    """Most recent served-but-not-submitted, non-swapped attempt, if any."""
    return conn.execute(
        """SELECT * FROM quiz_attempt
            WHERE session_id = ? AND student_token = ?
              AND submitted_at IS NULL AND swapped = 0
         ORDER BY served_at DESC LIMIT 1""",
        (session_id, student_token),
    ).fetchone()


def pick_next_item(conn, session_row, student_token):
    """Pick the next bank_item for this student. Respects exhaustion_policy."""
    session_id = session_row["id"]
    bank_id = session_row["bank_id"]
    tags = session_tags(session_row)
    if not tags:
        return None, "exhausted"

    placeholders = ",".join("?" for _ in tags)
    items = conn.execute(
        f"""SELECT * FROM bank_item
             WHERE bank_id = ? AND lecture_tag IN ({placeholders})
          ORDER BY id""",
        [bank_id, *tags],
    ).fetchall()
    if not items:
        return None, "exhausted"

    seen = seen_item_ids(conn, session_id, student_token)
    unseen = [i for i in items if i["id"] not in seen]

    if unseen:
        return secrets.choice(unseen), None

    if session_row["exhaustion_policy"] == "recycle":
        return secrets.choice(items), None

    return None, "exhausted"


def make_permutation(n_options):
    order = list(range(n_options))
    # Fisher–Yates with secrets.randbelow
    for i in range(n_options - 1, 0, -1):
        j = secrets.randbelow(i + 1)
        order[i], order[j] = order[j], order[i]
    return order


def serve_item(conn, session_row, student_token, item, counted):
    attempt_id = new_id()
    options = json.loads(item["options_json"])
    if session_row["permutation"] == "per_student":
        seed_src = f"{session_row['id']}|{student_token}|{item['id']}"
        seed = int(hashlib.sha256(seed_src.encode()).hexdigest(), 16)
        rng = random.Random(seed)
        order = list(range(len(options)))
        rng.shuffle(order)
    else:
        order = make_permutation(len(options))
    ord_value = None
    if counted:
        ord_value = count_ord(conn, session_row["id"], student_token) + 1
    conn.execute(
        """INSERT INTO quiz_attempt
             (id, session_id, student_token, bank_item_id, ord,
              option_order_json, served_at, swapped)
           VALUES (?, ?, ?, ?, ?, ?, ?, 0)""",
        (
            attempt_id,
            session_row["id"],
            student_token,
            item["id"],
            ord_value,
            json.dumps(order),
            now_ms(),
        ),
    )
    visible_options = [options[i] for i in order]
    return {
        "attempt_id": attempt_id,
        "stem": item["stem"],
        "options": visible_options,
        "ord": ord_value,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_xlsx(conn, session_id):
    session_row = session_by_id(conn, session_id)
    if not session_row:
        raise ValueError("Session not found.")

    students = conn.execute(
        """SELECT * FROM quiz_student WHERE session_id = ?
        ORDER BY student_number""",
        (session_id,),
    ).fetchall()

    wb = Workbook()
    summary = wb.active
    summary.title = "summary"
    summary.append([
        "student_number", "joined_at", "ended_at", "end_reason",
        "items_answered", "items_correct", "score_pct", "blur_count",
        "incident_count",
    ])

    detail = wb.create_sheet("detail")
    detail.append([
        "student_number", "ord", "bank_item_id", "lecture_tag", "stem",
        "chosen_option_text", "correct_option_text", "correct", "swapped",
        "served_at", "submitted_at", "response_ms",
    ])

    incidents = wb.create_sheet("incidents")
    incidents.append([
        "student_number", "event_type", "attempt_id", "client_ts",
        "server_ts", "metadata_json",
    ])

    for s in students:
        answered = count_answered(conn, session_id, s["student_token"])
        correct = count_correct(conn, session_id, s["student_token"])
        blur = count_blur(conn, session_id, s["student_token"])
        incident_count = count_incidents(conn, session_id, s["student_token"])
        score_pct = round(100.0 * correct / answered, 2) if answered else 0.0

        summary.append([
            s["student_number"],
            _epoch_to_dt(s["joined_at"]),
            _epoch_to_dt(s["ended_at"]),
            s["end_reason"] or "",
            answered,
            correct,
            score_pct,
            blur,
            incident_count,
        ])

        attempts = conn.execute(
            """SELECT a.*, b.lecture_tag AS item_tag, b.stem AS item_stem,
                      b.options_json AS item_options,
                      b.correct_index AS item_correct_index
                 FROM quiz_attempt a
            LEFT JOIN bank_item b ON b.id = a.bank_item_id
                WHERE a.session_id = ? AND a.student_token = ?
             ORDER BY a.served_at""",
            (session_id, s["student_token"]),
        ).fetchall()

        for a in attempts:
            options = json.loads(a["item_options"]) if a["item_options"] else []
            correct_text = (
                options[a["item_correct_index"]]
                if a["item_correct_index"] is not None and a["item_correct_index"] < len(options)
                else ""
            )
            chosen_text = (
                options[a["chosen_index"]]
                if a["chosen_index"] is not None and a["chosen_index"] < len(options)
                else ""
            )
            response_ms = None
            if a["submitted_at"] and a["served_at"]:
                response_ms = a["submitted_at"] - a["served_at"]

            detail.append([
                s["student_number"],
                a["ord"],
                a["bank_item_id"],
                a["item_tag"] or "",
                a["item_stem"] or "",
                chosen_text,
                correct_text,
                1 if a["correct"] == 1 else (0 if a["correct"] == 0 else ""),
                1 if a["swapped"] else 0,
                _epoch_to_dt(a["served_at"]),
                _epoch_to_dt(a["submitted_at"]),
                response_ms if response_ms is not None else "",
            ])

        incident_rows = conn.execute(
            """SELECT * FROM quiz_incident
                WHERE session_id = ? AND student_token = ?
             ORDER BY server_ts""",
            (session_id, s["student_token"]),
        ).fetchall()
        for incident in incident_rows:
            incidents.append([
                s["student_number"],
                incident["event_type"],
                incident["attempt_id"] or "",
                _epoch_to_dt(incident["client_ts"]),
                _epoch_to_dt(incident["server_ts"]),
                incident["metadata_json"] or "{}",
            ])

    # Freeze header rows; widen a few key columns.
    summary.freeze_panes = "A2"
    detail.freeze_panes = "A2"
    incidents.freeze_panes = "A2"
    for col_letter in ("A", "B", "C", "D"):
        summary.column_dimensions[col_letter].width = 18
    for col_letter, width in (("A", 16), ("E", 50), ("F", 30), ("G", 30)):
        detail.column_dimensions[col_letter].width = width
    for col_letter, width in (("A", 16), ("B", 24), ("C", 34), ("F", 70)):
        incidents.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _epoch_to_dt(epoch_ms):
    """Epoch ms → naïve local-time datetime for Excel.

    Excel datetimes carry no timezone, so we render in the server's local
    time (i.e. the instructor's laptop time when deployed per README).
    Don't switch to UTC here — instructors comparing the export to what
    they saw live during class would get a confusing offset.
    """
    if not epoch_ms:
        return ""
    return datetime.fromtimestamp(epoch_ms / 1000).replace(microsecond=0)


# ---------------------------------------------------------------------------
# Per-session locks
# ---------------------------------------------------------------------------


class SessionLocks:
    def __init__(self):
        self._lock = threading.Lock()
        self._locks = {}

    def get(self, session_id):
        with self._lock:
            lock = self._locks.get(session_id)
            if lock is None:
                lock = threading.Lock()
                self._locks[session_id] = lock
            return lock


SESSION_LOCKS = SessionLocks()


# ---------------------------------------------------------------------------
# HTTP handler
# ---------------------------------------------------------------------------


class Handler(BaseHTTPRequestHandler):
    server_version = "SessionQuiz/1.0"

    # -- Framework helpers --------------------------------------------------

    def _send_json(self, status, payload, extra_headers=None):
        body = json_dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        for name, value in (extra_headers or {}).items():
            self.send_header(name, value)
        self._send_cors()
        self.end_headers()
        self.wfile.write(body)

    def _send_binary(self, status, data, content_type, extra_headers=None):
        self.send_response(status)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        for name, value in (extra_headers or {}).items():
            self.send_header(name, value)
        self._send_cors()
        self.end_headers()
        self.wfile.write(data)

    def _send_error(self, status, message):
        self._send_json(status, {"error": message})

    def _send_cors(self):
        origin = self.server.allowed_origin or "*"
        self.send_header("Access-Control-Allow-Origin", origin)
        self.send_header("Access-Control-Allow-Methods",
                         "GET, POST, PUT, DELETE, OPTIONS")
        self.send_header("Access-Control-Allow-Headers",
                         "Content-Type, X-Admin-Key, X-Admin-Key-B64")
        self.send_header("Vary", "Origin")

    def log_message(self, fmt, *args):
        return

    def _parse_path(self):
        parsed = urlparse(self.path)
        return parsed.path, parse_qs(parsed.query)

    def _read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return b""
        return self.rfile.read(length)

    def _read_multipart(self):
        """Very small multipart/form-data parser for single-file uploads.

        Returns (fields, files) where fields is dict[str, str] and files is
        dict[str, {filename, content}]. Only supports a flat form.
        """
        ctype = self.headers.get("Content-Type", "")
        m = re.match(r"multipart/form-data;\s*boundary=(.+)", ctype, re.I)
        if not m:
            raise ValueError("Expected multipart/form-data.")
        boundary = m.group(1).strip().strip('"')
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            raise ValueError("Empty upload.")
        data = self.rfile.read(length)

        delim = ("--" + boundary).encode()
        parts = data.split(delim)
        fields, files = {}, {}
        for part in parts:
            if not part or part in (b"--\r\n", b"--"):
                continue
            if part.startswith(b"\r\n"):
                part = part[2:]
            if part.endswith(b"\r\n"):
                part = part[:-2]
            try:
                header_blob, body = part.split(b"\r\n\r\n", 1)
            except ValueError:
                continue
            headers_text = header_blob.decode("utf-8", errors="replace")
            disp_match = re.search(
                r'Content-Disposition:\s*form-data;\s*name="([^"]+)"(?:;\s*filename="([^"]*)")?',
                headers_text, re.I,
            )
            if not disp_match:
                continue
            name = disp_match.group(1)
            filename = disp_match.group(2)
            if filename:
                files[name] = {"filename": filename, "content": body}
            else:
                fields[name] = body.decode("utf-8", errors="replace")
        return fields, files

    def _read_admin_key(self):
        key_b64 = self.headers.get("X-Admin-Key-B64", "").strip()
        if key_b64:
            try:
                padding = "=" * (-len(key_b64) % 4)
                decoded = base64.b64decode(key_b64 + padding, validate=True)
                return decoded.decode("utf-8").strip()
            except (ValueError, UnicodeDecodeError):
                return ""
        return self.headers.get("X-Admin-Key", "").strip()

    def _check_admin(self):
        key = self._read_admin_key()
        expected = self.server.admin_key
        if not expected:
            return True  # dev/open mode
        if not key:
            return False
        return hmac.compare_digest(key, expected)

    def _require_admin(self):
        if not self._check_admin():
            self._send_error(401, "Valid admin key required.")
            return False
        return True

    def _conn(self):
        return self.server.get_conn()

    def _try_serve_static(self, path):
        if path in ("", "/"):
            target = "index.html"
        elif path == "/admin":
            target = "admin.html"
        else:
            target = path.lstrip("/")
        # Reject path traversal.
        if ".." in target.split("/"):
            return False
        candidate = (self.server.frontend_dir / target).resolve()
        try:
            candidate.relative_to(self.server.frontend_dir.resolve())
        except ValueError:
            return False
        if not candidate.is_file():
            return False
        ctype = _content_type_for(candidate.name)
        try:
            data = candidate.read_bytes()
        except OSError:
            return False
        # Static files are public; reuse the binary helper.
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self._send_cors()
        self.end_headers()
        self.wfile.write(data)
        return True

    # -- Routing ------------------------------------------------------------

    def do_OPTIONS(self):
        self.send_response(204)
        self._send_cors()
        self.end_headers()

    def do_GET(self):
        path, qs = self._parse_path()
        try:
            if path == "/quiz/health":
                return self._send_json(200, {"ok": True})
            # Static frontend — only when a frontend dir is configured.
            if self.server.frontend_dir and not path.startswith("/quiz/"):
                served = self._try_serve_static(path)
                if served:
                    return
            if path == "/quiz/next":
                return self._handle_next(qs)
            if path == "/quiz/live_status":
                return self._handle_live_status(qs)
            if path == "/quiz/admin/bank/list":
                if not self._require_admin():
                    return
                return self._handle_bank_list()
            if path == "/quiz/admin/bank/items":
                if not self._require_admin():
                    return
                return self._handle_bank_items(qs)
            if path == "/quiz/admin/session/list":
                if not self._require_admin():
                    return
                return self._handle_session_list()
            if path == "/quiz/admin/session/live":
                if not self._require_admin():
                    return
                return self._handle_session_live(qs)
            if path == "/quiz/admin/session/export":
                if not self._require_admin():
                    return
                return self._handle_export(qs)
            if path == "/quiz/admin/session/qr":
                if not self._require_admin():
                    return
                return self._handle_qr(qs)
            return self._send_error(404, "Not found.")
        except ValueError as exc:
            return self._send_error(400, str(exc))
        except Exception as exc:
            return self._send_error(500, f"Server error: {exc}")

    def do_POST(self):
        path, _qs = self._parse_path()
        try:
            if path == "/quiz/admin/bank/upload":
                if not self._require_admin():
                    return
                return self._handle_bank_upload()
            body = self._read_body()
            payload = parse_json(body)
            if path == "/quiz/join":
                return self._handle_join(payload)
            if path == "/quiz/answer":
                return self._handle_answer(payload)
            if path == "/quiz/blur":
                return self._handle_blur(payload)
            if path == "/quiz/incident":
                return self._handle_incident(payload)
            if path == "/quiz/admin/session/create":
                if not self._require_admin():
                    return
                return self._handle_session_create(payload)
            if path == "/quiz/admin/session/start":
                if not self._require_admin():
                    return
                return self._handle_session_start(payload)
            if path == "/quiz/admin/session/close":
                if not self._require_admin():
                    return
                return self._handle_session_close(payload)
            return self._send_error(404, "Not found.")
        except ValueError as exc:
            return self._send_error(400, str(exc))
        except Exception as exc:
            return self._send_error(500, f"Server error: {exc}")

    def do_DELETE(self):
        path, _qs = self._parse_path()
        try:
            if not self._require_admin():
                return
            m = re.match(r"^/quiz/admin/bank/([^/]+)$", path)
            if m:
                return self._handle_bank_delete(m.group(1))
            m = re.match(r"^/quiz/admin/session/([^/]+)$", path)
            if m:
                return self._handle_session_delete(m.group(1))
            return self._send_error(404, "Not found.")
        except ValueError as exc:
            return self._send_error(400, str(exc))
        except Exception as exc:
            return self._send_error(500, f"Server error: {exc}")

    # -- Admin: banks -------------------------------------------------------

    def _handle_bank_upload(self):
        fields, files = self._read_multipart()
        if "file" not in files:
            raise ValueError("No file part named 'file'.")
        upload = files["file"]
        given_name = (fields.get("name") or upload["filename"] or "bank").strip()

        bank_id = new_id()
        target_path = self.server.banks_dir / f"{bank_id}.xlsx"
        target_path.write_bytes(upload["content"])

        try:
            items, warnings, fatal = bank_io.read_xlsx(target_path)
        except bank_io.BankReadError as exc:
            target_path.unlink(missing_ok=True)
            raise ValueError(str(exc))

        if fatal:
            target_path.unlink(missing_ok=True)
            return self._send_json(
                400,
                {"error": "Bank has fatal row errors.", "fatal": fatal,
                 "warnings": warnings},
            )

        conn = self._conn()
        with conn:
            conn.execute(
                "INSERT INTO bank (id, name, uploaded_at, source_filename) "
                "VALUES (?, ?, ?, ?)",
                (bank_id, given_name, now_ms(), upload["filename"]),
            )
            for item in items:
                conn.execute(
                    """INSERT INTO bank_item
                         (id, bank_id, lecture_tag, stem, options_json,
                          correct_index, explanation)
                       VALUES (?, ?, ?, ?, ?, ?, ?)""",
                    (
                        new_id(), bank_id, item["lecture_tag"], item["stem"],
                        json.dumps(item["options"], ensure_ascii=False),
                        item["correct_index"], item["explanation"],
                    ),
                )

        tags = sorted({i["lecture_tag"] for i in items})
        tag_counts = {}
        for i in items:
            tag_counts[i["lecture_tag"]] = tag_counts.get(i["lecture_tag"], 0) + 1

        return self._send_json(201, {
            "bank_id": bank_id,
            "name": given_name,
            "item_count": len(items),
            "tags": tags,
            "tag_counts": tag_counts,
            "warnings": warnings,
        })

    def _handle_bank_list(self):
        conn = self._conn()
        rows = conn.execute(
            "SELECT * FROM bank ORDER BY uploaded_at DESC"
        ).fetchall()
        result = []
        for b in rows:
            items = conn.execute(
                "SELECT lecture_tag FROM bank_item WHERE bank_id = ?",
                (b["id"],),
            ).fetchall()
            tag_counts = {}
            for i in items:
                tag_counts[i["lecture_tag"]] = tag_counts.get(i["lecture_tag"], 0) + 1
            result.append({
                "bank_id": b["id"],
                "name": b["name"],
                "uploaded_at": b["uploaded_at"],
                "source_filename": b["source_filename"],
                "item_count": len(items),
                "tags": sorted(tag_counts.keys()),
                "tag_counts": tag_counts,
            })
        return self._send_json(200, {"banks": result})

    def _handle_bank_items(self, qs):
        bank_id = (qs.get("bank_id", [""])[0] or "").strip()
        if not bank_id:
            raise ValueError("bank_id is required.")
        conn = self._conn()
        rows = conn.execute(
            "SELECT lecture_tag FROM bank_item WHERE bank_id = ?",
            (bank_id,),
        ).fetchall()
        tag_counts = {}
        for r in rows:
            tag_counts[r["lecture_tag"]] = tag_counts.get(r["lecture_tag"], 0) + 1
        return self._send_json(200, {
            "bank_id": bank_id,
            "tags": sorted(tag_counts.keys()),
            "tag_counts": tag_counts,
            "item_count": len(rows),
        })

    def _handle_bank_delete(self, bank_id):
        conn = self._conn()
        row = conn.execute("SELECT id FROM bank WHERE id = ?", (bank_id,)).fetchone()
        if not row:
            return self._send_error(404, "Bank not found.")
        in_use = conn.execute(
            "SELECT COUNT(*) AS n FROM quiz_session WHERE bank_id = ?",
            (bank_id,),
        ).fetchone()["n"]
        if in_use:
            raise ValueError(
                "This bank is used by one or more sessions. Close or delete those first."
            )
        with conn:
            conn.execute("DELETE FROM bank_item WHERE bank_id = ?", (bank_id,))
            conn.execute("DELETE FROM bank WHERE id = ?", (bank_id,))
        (self.server.banks_dir / f"{bank_id}.xlsx").unlink(missing_ok=True)
        return self._send_json(200, {"ok": True})

    # -- Admin: sessions ----------------------------------------------------

    def _handle_session_create(self, payload):
        bank_id = (payload.get("bank_id") or "").strip()
        tags = requested_tags(payload)
        if not bank_id:
            raise ValueError("bank_id is required.")
        if not tags:
            raise ValueError("At least one lecture_tag is required.")

        conn = self._conn()
        bank = conn.execute("SELECT * FROM bank WHERE id = ?", (bank_id,)).fetchone()
        if not bank:
            raise ValueError("Unknown bank.")
        placeholders = ",".join("?" for _ in tags)
        tag_rows = conn.execute(
            f"""SELECT lecture_tag, COUNT(*) AS n
                  FROM bank_item
                 WHERE bank_id = ? AND lecture_tag IN ({placeholders})
              GROUP BY lecture_tag""",
            [bank_id, *tags],
        ).fetchall()
        tag_counts = {r["lecture_tag"]: r["n"] for r in tag_rows}
        missing_tags = [tag for tag in tags if tag_counts.get(tag, 0) == 0]
        if missing_tags:
            raise ValueError(
                "No items with selected tag(s): " + ", ".join(missing_tags)
            )
        available_item_count = sum(tag_counts[tag] for tag in tags)

        default_display = session_tag_label(tags)
        if len(default_display) > 80:
            default_display = f"{len(tags)} lecture_tags"
        display_name = (payload.get("display_name") or default_display).strip()
        if not display_name:
            display_name = default_display
        if len(display_name) > 80:
            raise ValueError("display_name must be at most 80 characters.")

        item_count = clamp_int(
            payload.get("item_count"), "item_count",
            *ITEM_COUNT_RANGE, DEFAULTS["item_count"],
        )
        duration = clamp_int(
            payload.get("duration_minutes"), "duration_minutes",
            *DURATION_RANGE, DEFAULTS["duration_minutes"],
        )
        swap = pick_enum(
            payload.get("swap_policy"), "swap_policy",
            VALID["swap_policy"], DEFAULTS["swap_policy"],
        )
        perm = pick_enum(
            payload.get("permutation"), "permutation",
            VALID["permutation"], DEFAULTS["permutation"],
        )
        feedback = pick_enum(
            payload.get("feedback"), "feedback",
            VALID["feedback"], DEFAULTS["feedback"],
        )
        exhaustion = pick_enum(
            payload.get("exhaustion_policy"), "exhaustion_policy",
            VALID["exhaustion_policy"], DEFAULTS["exhaustion_policy"],
        )
        security = pick_enum(
            payload.get("security_mode"), "security_mode",
            VALID["security_mode"], DEFAULTS["security_mode"],
        )

        # Generate a unique join code (retry a few times on rare collision).
        for _ in range(8):
            code = make_session_code()
            if not session_by_code(conn, code):
                break
        else:
            raise ValueError("Could not allocate a join code, try again.")

        session_id = new_id()
        pseudonym_key = secrets.token_bytes(PSEUDONYM_KEY_BYTES)
        lecture_tag_value = encode_session_tags(tags)
        with conn:
            conn.execute(
                """INSERT INTO quiz_session
                     (id, join_code, bank_id, lecture_tag, display_name,
                      item_count, duration_minutes, swap_policy, permutation,
                      feedback, exhaustion_policy, security_mode, pseudonym_key,
                      created_at, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'lobby')""",
                (session_id, code, bank_id, lecture_tag_value, display_name,
                 item_count, duration, swap, perm, feedback, exhaustion,
                 security, pseudonym_key, now_ms()),
            )

        return self._send_json(201, {
            "session_id": session_id,
            "join_code": code,
            "qr_png_url": f"/quiz/admin/session/qr?session_id={session_id}",
            "display_name": display_name,
            "bank_id": bank_id,
            "lecture_tag": session_tag_label(tags),
            "lecture_tags": tags,
            "available_item_count": available_item_count,
            "item_count": item_count,
            "duration_minutes": duration,
            "swap_policy": swap,
            "permutation": perm,
            "feedback": feedback,
            "exhaustion_policy": exhaustion,
            "security_mode": security,
            "status": "lobby",
        })

    def _handle_session_start(self, payload):
        session_id = (payload.get("session_id") or "").strip()
        if not session_id:
            raise ValueError("session_id is required.")
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")
        if row["status"] == "closed":
            raise ValueError("Session is already closed.")
        if row["status"] == "live":
            return self._send_json(200, {"started_at": row["started_at"]})
        lock = SESSION_LOCKS.get(session_id)
        with lock, conn:
            conn.execute(
                "UPDATE quiz_session SET status = 'live', started_at = ? WHERE id = ?",
                (now_ms(), session_id),
            )
        started = conn.execute(
            "SELECT started_at FROM quiz_session WHERE id = ?", (session_id,)
        ).fetchone()["started_at"]
        return self._send_json(200, {"started_at": started})

    def _handle_session_close(self, payload):
        session_id = (payload.get("session_id") or "").strip()
        if not session_id:
            raise ValueError("session_id is required.")
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")
        lock = SESSION_LOCKS.get(session_id)
        ts = now_ms()
        with lock, conn:
            if row["status"] != "closed":
                conn.execute(
                    "UPDATE quiz_session SET status='closed', closed_at=? WHERE id=?",
                    (ts, session_id),
                )
            conn.execute(
                """UPDATE quiz_student SET ended_at = ?, end_reason = 'instructor_closed'
                    WHERE session_id = ? AND ended_at IS NULL""",
                (ts, session_id),
            )
        return self._send_json(200, {"closed_at": ts})

    def _handle_session_delete(self, session_id):
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")
        with conn:
            conn.execute("DELETE FROM quiz_attempt WHERE session_id = ?", (session_id,))
            conn.execute("DELETE FROM quiz_student WHERE session_id = ?", (session_id,))
            conn.execute("DELETE FROM quiz_session WHERE id = ?", (session_id,))
        return self._send_json(200, {"ok": True})

    def _handle_session_list(self):
        conn = self._conn()
        rows = conn.execute(
            "SELECT * FROM quiz_session ORDER BY created_at DESC"
        ).fetchall()
        out = []
        for r in rows:
            n_students = conn.execute(
                "SELECT COUNT(*) AS n FROM quiz_student WHERE session_id = ?",
                (r["id"],),
            ).fetchone()["n"]
            out.append({**_session_public(r), "students": n_students})
        return self._send_json(200, {"sessions": out})

    def _handle_session_live(self, qs):
        session_id = (qs.get("session_id", [""])[0] or "").strip()
        if not session_id:
            raise ValueError("session_id is required.")
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")

        students = conn.execute(
            "SELECT * FROM quiz_student WHERE session_id = ? ORDER BY joined_at",
            (session_id,),
        ).fetchall()

        out_students = []
        for s in students:
            out_students.append({
                "student_number": s["student_number"],
                "answered": count_answered(conn, session_id, s["student_token"]),
                "current_ord": count_ord(conn, session_id, s["student_token"]),
                "swapped": count_blur(conn, session_id, s["student_token"]),
                "incidents": count_incidents(conn, session_id, s["student_token"]),
                "ended": bool(s["ended_at"]),
                "end_reason": s["end_reason"],
                "joined_at": s["joined_at"],
            })

        elapsed = (now_ms() - row["started_at"]) if row["started_at"] else 0
        remaining = compute_remaining_ms(row)
        return self._send_json(200, {
            "session": _session_public(row),
            "elapsed_ms": elapsed,
            "remaining_ms": remaining,
            "students": out_students,
        })

    def _handle_export(self, qs):
        session_id = (qs.get("session_id", [""])[0] or "").strip()
        if not session_id:
            raise ValueError("session_id is required.")
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")
        data = export_xlsx(conn, session_id)
        day = datetime.fromtimestamp(
            (row["created_at"] or now_ms()) / 1000
        ).strftime("%Y%m%d")
        filename = f"quiz_{row['join_code']}_{day}.xlsx"
        return self._send_binary(
            200, data,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            extra_headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            },
        )

    def _handle_qr(self, qs):
        session_id = (qs.get("session_id", [""])[0] or "").strip()
        if not session_id:
            raise ValueError("session_id is required.")
        conn = self._conn()
        row = session_by_id(conn, session_id)
        if not row:
            return self._send_error(404, "Session not found.")
        url = self.server.public_join_url(row["join_code"])
        png = qrmod.png_bytes(url)
        return self._send_binary(200, png, "image/png")

    # -- Student: join / next / answer / blur / live_status ----------------

    def _handle_join(self, payload):
        code = (payload.get("code") or "").strip().upper()
        if not code:
            raise ValueError("code is required.")
        student_number = normalize_student_number(payload.get("student_number") or "")

        conn = self._conn()
        row = session_by_code(conn, code)
        if not row:
            raise ValueError("Unknown session code.")
        if row["status"] == "closed":
            raise ValueError("This session has ended.")

        token = derive_student_token(row["pseudonym_key"], student_number)
        existing = student_row(conn, row["id"], token)
        if not existing:
            with conn:
                conn.execute(
                    """INSERT INTO quiz_student
                         (session_id, student_token, student_number, joined_at)
                       VALUES (?, ?, ?, ?)""",
                    (row["id"], token, student_number, now_ms()),
                )

        return self._send_json(200, {
            "student_token": token,
            "session_id": row["id"],
            "join_code": row["join_code"],
            "status": row["status"],
            "display_name": row["display_name"],
            "rules_text": rules_text(row),
            "item_count": row["item_count"],
            "duration_minutes": row["duration_minutes"],
            "feedback": row["feedback"],
            "swap_policy": row["swap_policy"],
            "security_mode": row["security_mode"],
        })

    def _handle_live_status(self, qs):
        token = (qs.get("student_token", [""])[0] or "").strip()
        if not token:
            raise ValueError("student_token is required.")
        conn = self._conn()
        s = conn.execute(
            """SELECT st.*, qs.status AS session_status, qs.display_name,
                      qs.item_count, qs.started_at, qs.duration_minutes,
                      qs.security_mode
                 FROM quiz_student st
                 JOIN quiz_session qs ON qs.id = st.session_id
                WHERE st.student_token = ?""",
            (token,),
        ).fetchone()
        if not s:
            raise ValueError("Unknown student token.")
        payload = {
            "status": s["session_status"],
            "display_name": s["display_name"],
            "item_count": s["item_count"],
            "security_mode": s["security_mode"],
            "ended": bool(s["ended_at"]),
            "end_reason": s["end_reason"],
        }
        if s["started_at"]:
            payload["remaining_ms"] = max(
                0,
                s["started_at"] + s["duration_minutes"] * 60_000 - now_ms(),
            )
        return self._send_json(200, payload)

    def _handle_incident(self, payload):
        token = (payload.get("student_token") or "").strip()
        attempt_id = (payload.get("attempt_id") or "").strip() or None
        event_type = (payload.get("event_type") or "").strip().lower()
        if not token:
            raise ValueError("student_token is required.")
        if not event_type:
            raise ValueError("event_type is required.")
        if not re.match(r"^[a-z0-9_:-]{1,64}$", event_type):
            raise ValueError("event_type must be 1-64 lowercase identifier characters.")

        client_ts = payload.get("client_ts")
        if client_ts in ("", None):
            client_ts = None
        else:
            try:
                client_ts = int(client_ts)
            except (TypeError, ValueError):
                raise ValueError("client_ts must be an integer epoch ms.")

        metadata = payload.get("metadata") or {}
        if not isinstance(metadata, dict):
            raise ValueError("metadata must be an object.")
        metadata_json = json_dumps(metadata)
        if len(metadata_json.encode("utf-8")) > 4096:
            raise ValueError("metadata is too large.")

        conn = self._conn()
        if attempt_id:
            attempt = conn.execute(
                "SELECT * FROM quiz_attempt WHERE id = ?", (attempt_id,)
            ).fetchone()
            if not attempt:
                raise ValueError("Unknown attempt.")
            if attempt["student_token"] != token:
                raise ValueError("Attempt does not belong to this student.")
            session_id = attempt["session_id"]
        else:
            student = student_row_by_token(conn, token)
            if not student:
                raise ValueError("Unknown student token.")
            session_id = student["session_id"]

        with conn:
            conn.execute(
                """INSERT INTO quiz_incident
                     (id, session_id, student_token, attempt_id, event_type,
                      client_ts, server_ts, metadata_json)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    new_id(), session_id, token, attempt_id, event_type,
                    client_ts, now_ms(), metadata_json,
                ),
            )
        return self._send_json(201, {"ok": True})

    def _handle_next(self, qs):
        token = (qs.get("student_token", [""])[0] or "").strip()
        if not token:
            raise ValueError("student_token is required.")
        conn = self._conn()
        s = student_row_by_token(conn, token)
        if not s:
            raise ValueError("Unknown student token.")
        session_row = session_by_id(conn, s["session_id"])
        if not session_row:
            return self._send_error(404, "Session not found.")

        # Already ended?
        if s["ended_at"]:
            return self._send_json(200, _ended_payload(
                conn, session_row, s, s["end_reason"]
            ))

        # If session not live, nothing to serve.
        if session_row["status"] == "lobby":
            return self._send_json(200, {"session_status": "lobby"})
        if session_row["status"] == "closed":
            with conn:
                end_student(conn, session_row["id"], token, "instructor_closed")
            s = student_row_by_token(conn, token)
            return self._send_json(200, _ended_payload(
                conn, session_row, s, s["end_reason"] or "instructor_closed"
            ))

        lock = SESSION_LOCKS.get(session_row["id"])
        with lock, conn:
            # Re-read student inside the lock to catch races.
            s = student_row_by_token(conn, token)
            session_row = session_by_id(conn, s["session_id"])

            if session_row["status"] == "lobby":
                return self._send_json(200, {"session_status": "lobby"})

            reason = ending_reason_for_student(conn, session_row, s)
            if reason:
                s = student_row_by_token(conn, token)
                return self._send_json(200, _ended_payload(
                    conn, session_row, s, reason
                ))

            # Reuse in-flight attempt if any.
            pending = open_attempt(conn, session_row["id"], token)
            if pending:
                item = conn.execute(
                    "SELECT * FROM bank_item WHERE id = ?",
                    (pending["bank_item_id"],),
                ).fetchone()
                options = json.loads(item["options_json"])
                order = json.loads(pending["option_order_json"])
                return self._send_json(200, {
                    "attempt_id": pending["id"],
                    "stem": item["stem"],
                    "options": [options[i] for i in order],
                    "ord": pending["ord"],
                    "item_count": session_row["item_count"],
                    "security_mode": session_row["security_mode"],
                    "remaining_ms": compute_remaining_ms(session_row),
                })

            answered = count_answered(conn, session_row["id"], token)
            if answered >= session_row["item_count"]:
                end_student(conn, session_row["id"], token, "completed")
                s = student_row_by_token(conn, token)
                return self._send_json(200, _ended_payload(
                    conn, session_row, s, "completed"
                ))

            item, reason = pick_next_item(conn, session_row, token)
            if not item:
                end_student(conn, session_row["id"], token, reason or "exhausted")
                s = student_row_by_token(conn, token)
                return self._send_json(200, _ended_payload(
                    conn, session_row, s, reason or "exhausted"
                ))

            served = serve_item(conn, session_row, token, item, counted=True)

        return self._send_json(200, {
            **served,
            "item_count": session_row["item_count"],
            "security_mode": session_row["security_mode"],
            "remaining_ms": compute_remaining_ms(session_row),
        })

    def _handle_answer(self, payload):
        attempt_id = (payload.get("attempt_id") or "").strip()
        if not attempt_id:
            raise ValueError("attempt_id is required.")
        try:
            chosen_visible_index = int(payload.get("chosen_visible_index"))
        except (TypeError, ValueError):
            raise ValueError("chosen_visible_index must be an integer.")

        conn = self._conn()
        row = conn.execute(
            "SELECT * FROM quiz_attempt WHERE id = ?", (attempt_id,)
        ).fetchone()
        if not row:
            raise ValueError("Unknown attempt.")
        lock = SESSION_LOCKS.get(row["session_id"])
        with lock, conn:
            row = conn.execute(
                "SELECT * FROM quiz_attempt WHERE id = ?", (attempt_id,)
            ).fetchone()
            if not row:
                raise ValueError("Unknown attempt.")
            if row["submitted_at"] is not None:
                raise ValueError("Attempt already submitted.")
            if row["swapped"]:
                raise ValueError("This attempt was swapped.")

            session_row = session_by_id(conn, row["session_id"])
            if not session_row:
                return self._send_error(404, "Session not found.")
            student = student_row(conn, row["session_id"], row["student_token"])
            if not student:
                raise ValueError("Unknown student token.")
            if session_row["status"] == "lobby":
                return self._send_json(
                    409,
                    {"error": "Session is not live."},
                )
            reason = ending_reason_for_student(conn, session_row, student)
            if reason:
                student = student_row(conn, row["session_id"], row["student_token"])
                return self._send_json(
                    200,
                    _ended_payload(conn, session_row, student, reason),
                )

            item = conn.execute(
                "SELECT * FROM bank_item WHERE id = ?", (row["bank_item_id"],)
            ).fetchone()
            order = json.loads(row["option_order_json"])
            if chosen_visible_index < 0 or chosen_visible_index >= len(order):
                raise ValueError("chosen_visible_index out of range.")
            canonical_index = order[chosen_visible_index]
            correct = 1 if canonical_index == item["correct_index"] else 0
            correct_visible_index = order.index(item["correct_index"])

            # Conditional update guards against a TOCTOU race with /quiz/blur.
            cur = conn.execute(
                """UPDATE quiz_attempt
                      SET chosen_index = ?, correct = ?, submitted_at = ?
                    WHERE id = ? AND submitted_at IS NULL AND swapped = 0""",
                (canonical_index, correct, now_ms(), attempt_id),
            )
            committed = cur.rowcount > 0

        if not committed:
            return self._send_json(
                409,
                {"error": "Attempt was swapped or already submitted."},
            )

        payload_out = {"correct": bool(correct)}
        if session_row["feedback"] == "immediate":
            payload_out["explanation"] = item["explanation"] or ""
            options_canon = json.loads(item["options_json"])
            payload_out["correct_option_text"] = options_canon[item["correct_index"]]
            payload_out["correct_visible_index"] = correct_visible_index
        return self._send_json(200, payload_out)

    def _handle_blur(self, payload):
        attempt_id = (payload.get("attempt_id") or "").strip()
        if not attempt_id:
            raise ValueError("attempt_id is required.")
        conn = self._conn()
        row = conn.execute(
            "SELECT * FROM quiz_attempt WHERE id = ?", (attempt_id,)
        ).fetchone()
        if not row:
            raise ValueError("Unknown attempt.")
        if row["submitted_at"] is not None:
            # Already submitted — blur is moot.
            return self._send_json(200, {"swapped": False, "submitted": True})
        if row["swapped"]:
            return self._send_json(200, {"swapped": True})

        session_row = session_by_id(conn, row["session_id"])
        token = row["student_token"]
        lock = SESSION_LOCKS.get(row["session_id"])
        with lock, conn:
            # Symmetric guard: if /quiz/answer committed between the read
            # above and this UPDATE, do nothing — the student earned that
            # answer, and a stray blur from the same client should not
            # rescind it.
            cur = conn.execute(
                """UPDATE quiz_attempt SET swapped = 1, ord = NULL
                    WHERE id = ? AND submitted_at IS NULL AND swapped = 0""",
                (attempt_id,),
            )
            swapped_now = cur.rowcount > 0
            if swapped_now and session_row["swap_policy"] == "hard":
                end_student(conn, session_row["id"], token, "blur_hard")

        if not swapped_now:
            # Race winner was the answer; treat as a no-op for the client.
            return self._send_json(200, {"swapped": False, "submitted": True})
        if session_row["swap_policy"] == "hard":
            return self._send_json(200, {
                "session_ended": True,
                "reason": "blur_hard",
            })
        return self._send_json(200, {"swapped": True})


# ---------------------------------------------------------------------------
# Helpers used by handlers
# ---------------------------------------------------------------------------


def student_row_by_token(conn, token):
    return conn.execute(
        "SELECT * FROM quiz_student WHERE student_token = ?", (token,)
    ).fetchone()


def _session_public(row):
    tags = session_tags(row)
    return {
        "session_id": row["id"],
        "join_code": row["join_code"],
        "bank_id": row["bank_id"],
        "lecture_tag": session_tag_label(tags),
        "lecture_tags": tags,
        "display_name": row["display_name"],
        "item_count": row["item_count"],
        "duration_minutes": row["duration_minutes"],
        "swap_policy": row["swap_policy"],
        "permutation": row["permutation"],
        "feedback": row["feedback"],
        "exhaustion_policy": row["exhaustion_policy"],
        "security_mode": row["security_mode"],
        "status": row["status"],
        "created_at": row["created_at"],
        "started_at": row["started_at"],
        "closed_at": row["closed_at"],
    }


def _ended_payload(conn, session_row, student, reason):
    answered = count_answered(conn, session_row["id"], student["student_token"])
    correct = count_correct(conn, session_row["id"], student["student_token"])
    return {
        "session_ended": True,
        "reason": reason or "instructor_closed",
        "score": {"answered": answered, "correct": correct,
                  "item_count": session_row["item_count"]},
    }


# ---------------------------------------------------------------------------
# Server wrapper
# ---------------------------------------------------------------------------


_STATIC_TYPES = {
    ".html": "text/html; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".json": "application/json; charset=utf-8",
    ".png": "image/png",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".svg": "image/svg+xml",
    ".ico": "image/x-icon",
    ".txt": "text/plain; charset=utf-8",
}


def _content_type_for(filename):
    for ext, ctype in _STATIC_TYPES.items():
        if filename.lower().endswith(ext):
            return ctype
    return "application/octet-stream"


class QuizServer(ThreadingHTTPServer):
    def __init__(self, address, handler, db_path, admin_key, allowed_origin,
                 banks_dir, public_base_url, frontend_dir):
        super().__init__(address, handler)
        self.db_path = db_path
        self.admin_key = admin_key
        self.allowed_origin = allowed_origin
        self.banks_dir = banks_dir
        self.public_base_url = public_base_url.rstrip("/")
        self.frontend_dir = frontend_dir
        self._conn_lock = threading.Lock()
        self._thread_conns = {}

    def get_conn(self):
        ident = threading.get_ident()
        with self._conn_lock:
            conn = self._thread_conns.get(ident)
            if conn is None:
                conn = dbmod.connect(self.db_path)
                self._thread_conns[ident] = conn
            return conn

    def public_join_url(self, code):
        base = self.public_base_url or ""
        return f"{base}/?code={code}" if base else f"/?code={code}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--host", default="0.0.0.0")
    parser.add_argument("--port", type=int, default=8789)
    parser.add_argument("--db", default=None)
    args = parser.parse_args()

    root = Path(__file__).parent
    db_path = args.db or os.environ.get(
        "QUIZ_DB_PATH", str(root / "data" / "quiz.db"),
    )
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)

    banks_dir = root / "banks"
    banks_dir.mkdir(parents=True, exist_ok=True)

    admin_key = os.environ.get("QUIZ_ADMIN_KEY", "").strip()
    allowed_origin = os.environ.get("QUIZ_ALLOWED_ORIGIN", "").strip() or None
    public_base_url = os.environ.get("QUIZ_PUBLIC_URL", "").strip()

    frontend_env = os.environ.get("QUIZ_FRONTEND_DIR", "").strip()
    if frontend_env:
        frontend_dir = Path(frontend_env)
    else:
        candidate = (root.parent / "frontend").resolve()
        frontend_dir = candidate if candidate.is_dir() else None

    dbmod.init_schema(dbmod.connect(db_path))

    server = QuizServer(
        (args.host, args.port), Handler, db_path, admin_key, allowed_origin,
        banks_dir, public_base_url, frontend_dir,
    )
    print(f"[quiz] listening on {args.host}:{args.port}")
    print(f"[quiz] db: {db_path}")
    print(f"[quiz] banks: {banks_dir}")
    if admin_key:
        print("[quiz] admin key is set (X-Admin-Key or X-Admin-Key-B64 header)")
    else:
        print("[quiz] WARNING: no admin key set (dev mode, all admin routes open)")
    if public_base_url:
        print(f"[quiz] public QR base URL: {public_base_url}")
    if frontend_dir:
        print(f"[quiz] serving frontend from {frontend_dir}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("[quiz] shutting down")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
