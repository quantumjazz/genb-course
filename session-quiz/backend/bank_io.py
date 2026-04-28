"""Read and validate question-bank .xlsx files.

Expected format (sheet `items`, header on row 1):

  lecture_tag | stem | option_a | option_b | option_c | option_d |
  option_e    | option_f | correct | explanation

See session-quiz-spec.md §8 for the contract. Row-level warnings do not
reject the bank; fatal errors (missing required cell, bad `correct`
letter) do.
"""

from openpyxl import load_workbook


OPTION_LETTERS = ["a", "b", "c", "d", "e", "f"]
REQUIRED_COLUMNS = {"lecture_tag", "stem", "option_a", "option_b", "correct"}
KNOWN_COLUMNS = {
    "lecture_tag", "stem", "option_a", "option_b", "option_c", "option_d",
    "option_e", "option_f", "correct", "explanation",
}


class BankReadError(Exception):
    """Fatal error — the bank cannot be loaded at all (bad sheet, header)."""


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def read_xlsx(path):
    """Return (items, warnings, fatal).

    - items: list of dicts with keys
        lecture_tag, stem, options (list of non-empty strings),
        correct_index (int), explanation (str, may be empty).
    - warnings: list of human-readable per-row notes.
    - fatal: list of fatal errors. When non-empty, `items` should not be
      used.
    """
    warnings = []
    fatal = []
    items = []

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise BankReadError(f"Could not open workbook: {exc}") from exc

    if "items" not in wb.sheetnames:
        raise BankReadError(
            "Workbook is missing a sheet named 'items'. Rename your sheet to "
            "'items' and re-upload."
        )

    ws = wb["items"]

    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        raise BankReadError("Sheet 'items' is empty.")

    headers = [
        (_cell_text(cell).lower() if cell is not None else "")
        for cell in header_row
    ]

    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise BankReadError(
            "Sheet 'items' is missing required columns: "
            + ", ".join(sorted(missing))
        )

    for unknown in set(headers) - KNOWN_COLUMNS:
        if unknown:
            warnings.append(f"Ignoring unknown column '{unknown}'.")

    index_of = {name: headers.index(name) for name in headers if name}

    for row_number, row in enumerate(rows, start=2):
        if row is None or all(cell is None or _cell_text(cell) == "" for cell in row):
            continue
        try:
            item = _parse_row(row, index_of, row_number)
        except ValueError as exc:
            fatal.append(f"Row {row_number}: {exc}")
            continue
        items.append(item)

    if not items and not fatal:
        fatal.append("No question rows found after the header.")

    return items, warnings, fatal


def _parse_row(row, index_of, row_number):
    def cell(name):
        idx = index_of.get(name)
        if idx is None or idx >= len(row):
            return ""
        return _cell_text(row[idx])

    lecture_tag = cell("lecture_tag")
    stem = cell("stem")
    correct_raw = cell("correct").lower()
    explanation = cell("explanation")

    if not lecture_tag:
        raise ValueError("lecture_tag is empty.")
    if not stem:
        raise ValueError("stem is empty.")
    if not correct_raw:
        raise ValueError("correct is empty.")

    option_texts = []
    option_letters_present = []
    for letter in OPTION_LETTERS:
        text = cell(f"option_{letter}")
        if text:
            option_texts.append(text)
            option_letters_present.append(letter)

    if len(option_texts) < 2:
        raise ValueError("at least two options are required.")

    if correct_raw not in option_letters_present:
        raise ValueError(
            f"correct='{correct_raw}' does not map to a filled option "
            f"(filled: {', '.join(option_letters_present)})."
        )

    correct_index = option_letters_present.index(correct_raw)

    return {
        "lecture_tag": lecture_tag,
        "stem": stem,
        "options": option_texts,
        "correct_index": correct_index,
        "explanation": explanation,
    }
